"""Format-specific text extraction and fallback orchestration."""
import csv
import logging
from pathlib import Path

from fileExtraction.capabilities import (
    DOC_AVAILABLE,
    DOCX_AVAILABLE,
    PDF_AVAILABLE,
    XLSX_AVAILABLE,
)
from fileExtraction.constants import SUPPORTED_EXTENSIONS
from fileExtraction.detection import detect_file_extension

logger = logging.getLogger(__name__)

if PDF_AVAILABLE:
    import pypdf
if DOCX_AVAILABLE:
    from docx import Document
if DOC_AVAILABLE:
    import docx2python
if XLSX_AVAILABLE:
    from openpyxl import load_workbook


def extract_pdf(file_path: str):
    if not PDF_AVAILABLE:
        return None, "PDF extraction library not available"
    try:
        text_content = []
        with open(file_path, "rb") as file:
            pdf_reader = pypdf.PdfReader(file)
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text:
                    text_content.append(text)
        return "\n".join(text_content), None
    except Exception as exc:
        logger.error("PDF extraction error: %s", exc)
        return None, str(exc)


def extract_docx(file_path: str):
    if not DOCX_AVAILABLE:
        return None, "DOCX extraction library not available"
    try:
        doc = Document(file_path)
        text_content = [paragraph.text for paragraph in doc.paragraphs if paragraph.text]
        return "\n".join(text_content), None
    except Exception as exc:
        logger.error("DOCX extraction error: %s", exc)
        return None, str(exc)


def extract_doc(file_path: str):
    if not DOC_AVAILABLE:
        return None, "DOC extraction library not available"
    try:
        doc_content = docx2python.docx2python(file_path)
        return doc_content.text, None
    except Exception as exc:
        logger.error("DOC extraction error: %s", exc)
        return None, str(exc)


def extract_csv(file_path: str):
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for encoding in encodings:
        try:
            content = []
            with open(file_path, "r", encoding=encoding) as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    if row:
                        content.append(",".join(str(cell) for cell in row))
            return "\n".join(content), None
        except (UnicodeDecodeError, csv.Error) as exc:
            logger.debug("CSV extraction with %s failed: %s", encoding, exc)
            continue
        except Exception as exc:
            logger.error("CSV extraction error: %s", exc)
            return None, str(exc)
    return None, "Could not parse CSV file with any supported encoding"


def extract_txt(file_path: str):
    encodings = ["utf-8", "latin-1", "cp1252"]
    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as file:
                return file.read(), None
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            logger.error("TXT extraction error: %s", exc)
            return None, str(exc)
    return None, "Could not read text file with any supported encoding"


def _read_xlsx_rows(workbook) -> list[str]:
    rows = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            if row and any(cell is not None and str(cell).strip() for cell in row):
                rows.append(",".join("" if cell is None else str(cell) for cell in row))
    return rows


def extract_xlsx(file_path: str):
    if not XLSX_AVAILABLE:
        return None, "XLSX extraction library not available"
    try:
        rows = []
        for data_only in (True, False):
            workbook = load_workbook(file_path, read_only=False, data_only=data_only)
            try:
                rows = _read_xlsx_rows(workbook)
            finally:
                workbook.close()
            if rows:
                break
        if not rows:
            return None, "No cell values found in XLSX file"
        return "\n".join(rows), None
    except Exception as exc:
        logger.error("XLSX extraction error: %s", exc)
        return None, str(exc)


EXTRACTION_FUNCTIONS = {
    ".pdf": extract_pdf,
    ".docx": extract_docx,
    ".doc": extract_doc,
    ".csv": extract_csv,
    ".txt": extract_txt,
    ".xlsx": extract_xlsx,
}


def try_extract_with_fallback(file_path: str, file_extension: str | None = None):
    """
    Try extraction with the known extension, or probe supported formats.

    Returns:
        (content, detected_extension, error_message)
    """
    if not file_extension:
        file_extension = detect_file_extension(file_path)

    if file_extension and file_extension in EXTRACTION_FUNCTIONS:
        extract_func = EXTRACTION_FUNCTIONS[file_extension]
        try:
            content, error = extract_func(file_path)
            if content and not error:
                return content, file_extension, None
            return None, file_extension, error or "Failed to extract content from file"
        except Exception as exc:
            logger.debug("Extraction with %s failed: %s", file_extension, exc)
            return None, file_extension, str(exc)

    for ext, extract_func in EXTRACTION_FUNCTIONS.items():
        if file_extension and ext == file_extension:
            continue
        try:
            content, error = extract_func(file_path)
            if content and not error:
                logger.info("Detected file type as %s via content analysis", ext)
                return content, ext, None
        except Exception as exc:
            logger.debug("Extraction with %s failed: %s", ext, exc)
            continue

    return None, file_extension, "Could not extract content with any supported method"
