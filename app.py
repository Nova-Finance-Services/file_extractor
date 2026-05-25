"""
File Extractor API with security enhancements
"""
from flask import Flask, request, jsonify
from functools import wraps
import requests
import os
import tempfile
import base64
import binascii
from pathlib import Path
import csv
import logging
import zipfile
from io import BytesIO
from urllib.parse import urlparse
from functools import lru_cache
from dotenv import load_dotenv
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuration
@lru_cache()
def get_config():
    """Load configuration from environment variables"""
    return {
        'MAX_FILE_SIZE': int(os.environ.get('MAX_FILE_SIZE', 50 * 1024 * 1024)),  # 50MB default
        'REQUEST_TIMEOUT': int(os.environ.get('REQUEST_TIMEOUT', 30)),
        'ALLOWED_SCHEMES': ['http', 'https'],
        'BLOCKED_HOSTS': ['localhost', '127.0.0.1', '0.0.0.0', '::1', '169.254.169.254'],  # AWS metadata
        'FILE_EXTRACTOR_KEY': os.environ.get('FILE_EXTRACTOR_KEY', ''),
    }

CONFIG = get_config()

# Initialize Flask app
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = CONFIG['MAX_FILE_SIZE'] * 2  # Allow larger responses

# Initialize rate limiter
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["100 per hour", "10 per minute"],
    storage_uri="memory://"
)

# PDF extraction
try:
    import pypdf
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("pypdf not available. PDF extraction disabled.")

# DOCX extraction
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    logger.warning("python-docx not available. DOCX extraction disabled.")

# DOC extraction (old format)
try:
    import docx2python
    DOC_AVAILABLE = True
except ImportError:
    DOC_AVAILABLE = False
    logger.warning("docx2python not available. DOC extraction disabled.")

# XLSX extraction
try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not available. XLSX extraction disabled.")

# Supported file types mapping
SUPPORTED_EXTENSIONS = ['.pdf', '.doc', '.docx', '.csv', '.txt', '.xlsx']

def require_api_key(f):
    """
    Decorator to require API key authentication
    
    Args:
        f: Function to protect
        
    Returns:
        Decorated function
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Get API key from environment
        api_key = CONFIG.get('FILE_EXTRACTOR_KEY', '')
        
        # If no API key is configured, skip authentication
        if not api_key:
            logger.warning("No API key configured. Authentication disabled.")
            return f(*args, **kwargs)
        
        # Get Authorization header
        auth_header = request.headers.get('Authorization', '')
        
        # Check if Authorization header is present
        if not auth_header:
            logger.warning("Missing Authorization header")
            return jsonify({
                'error': 'Missing Authorization header. Please provide API key in Authorization header.'
            }), 401
        
        # Extract API key from header (supports "Bearer <key>" or just "<key>")
        if auth_header.startswith('Bearer '):
            provided_key = auth_header[7:]  # Remove "Bearer " prefix
        else:
            provided_key = auth_header
        
        # Validate API key
        if provided_key != api_key:
            logger.warning(f"Invalid API key attempt from {request.remote_addr}")
            return jsonify({
                'error': 'Invalid API key'
            }), 401
        
        return f(*args, **kwargs)
    return decorated_function

def validate_url(url):
    """
    Validate URL to prevent SSRF attacks
    
    Args:
        url: URL string to validate
        
    Returns:
        tuple: (is_valid, error_message)
    """
    if not url or not isinstance(url, str):
        return False, "URL must be a non-empty string"
    
    if len(url) > 2048:
        return False, "URL too long (max 2048 characters)"
    
    try:
        parsed = urlparse(url)
        
        # Only allow http/https
        if parsed.scheme not in CONFIG['ALLOWED_SCHEMES']:
            return False, f"Only {', '.join(CONFIG['ALLOWED_SCHEMES'])} URLs are allowed"
        
        # Block internal/private IPs
        host = parsed.hostname
        if not host:
            return False, "Invalid URL: missing hostname"
        
        # Check for blocked hosts
        if host.lower() in CONFIG['BLOCKED_HOSTS']:
            return False, "Internal URLs are not allowed"
        
        # Block private IP ranges (basic check)
        host_lower = host.lower()
        if (host_lower.startswith('10.') or 
            host_lower.startswith('192.168.') or 
            host_lower.startswith('172.16.') or
            host_lower.startswith('172.17.') or
            host_lower.startswith('172.18.') or
            host_lower.startswith('172.19.') or
            host_lower.startswith('172.20.') or
            host_lower.startswith('172.21.') or
            host_lower.startswith('172.22.') or
            host_lower.startswith('172.23.') or
            host_lower.startswith('172.24.') or
            host_lower.startswith('172.25.') or
            host_lower.startswith('172.26.') or
            host_lower.startswith('172.27.') or
            host_lower.startswith('172.28.') or
            host_lower.startswith('172.29.') or
            host_lower.startswith('172.30.') or
            host_lower.startswith('172.31.')):
            return False, "Private IP addresses are not allowed"
        
        return True, None
    except Exception as e:
        logger.error(f"URL validation error: {str(e)}")
        return False, f"Invalid URL format: {str(e)}"

def extract_pdf(file_path):
    """Extract text from PDF file"""
    if not PDF_AVAILABLE:
        return None, "PDF extraction library not available"
    
    try:
        text_content = []
        with open(file_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text:
                    text_content.append(text)
        return '\n'.join(text_content), None
    except Exception as e:
        logger.error(f"PDF extraction error: {str(e)}")
        return None, str(e)

def extract_docx(file_path):
    """Extract text from DOCX file"""
    if not DOCX_AVAILABLE:
        return None, "DOCX extraction library not available"
    
    try:
        doc = Document(file_path)
        text_content = []
        for paragraph in doc.paragraphs:
            if paragraph.text:
                text_content.append(paragraph.text)
        return '\n'.join(text_content), None
    except Exception as e:
        logger.error(f"DOCX extraction error: {str(e)}")
        return None, str(e)

def extract_doc(file_path):
    """Extract text from DOC file (old format)"""
    if not DOC_AVAILABLE:
        return None, "DOC extraction library not available"
    
    try:
        doc_content = docx2python.docx2python(file_path)
        return doc_content.text, None
    except Exception as e:
        logger.error(f"DOC extraction error: {str(e)}")
        return None, str(e)

def extract_csv(file_path):
    """Extract content from CSV file with better error handling"""
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
    
    for encoding in encodings:
        try:
            content = []
            with open(file_path, 'r', encoding=encoding) as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    if row:  # Skip empty rows
                        content.append(','.join(str(cell) for cell in row))
            return '\n'.join(content), None
        except (UnicodeDecodeError, csv.Error) as e:
            logger.debug(f"CSV extraction with {encoding} failed: {str(e)}")
            continue
        except Exception as e:
            logger.error(f"CSV extraction error: {str(e)}")
            return None, str(e)
    
    return None, "Could not parse CSV file with any supported encoding"

def extract_txt(file_path):
    """Extract content from TXT file"""
    encodings = ['utf-8', 'latin-1', 'cp1252']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                return file.read(), None
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logger.error(f"TXT extraction error: {str(e)}")
            return None, str(e)
    
    return None, "Could not read text file with any supported encoding"

def is_xlsx_bytes(file_bytes):
    """Detect XLSX payloads by ZIP structure."""
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
            return any(name.startswith('xl/') for name in archive.namelist())
    except (zipfile.BadZipFile, OSError, ValueError):
        return False

def is_xlsx_file(file_path):
    """Detect XLSX files by ZIP structure (Office Open XML spreadsheet)."""
    try:
        if not zipfile.is_zipfile(file_path):
            return False
        with zipfile.ZipFile(file_path) as archive:
            return any(name.startswith('xl/') for name in archive.namelist())
    except (zipfile.BadZipFile, OSError):
        return False

def ensure_file_extension_path(file_path, file_extension):
    """Rename temp files so libraries that require extensions (e.g. openpyxl) can open them."""
    if not file_extension:
        return file_path
    current_suffix = Path(file_path).suffix.lower()
    if current_suffix == file_extension:
        return file_path
    new_path = str(Path(file_path).with_suffix(file_extension))
    os.rename(file_path, new_path)
    return new_path

def _read_xlsx_rows(workbook):
    """Read non-empty rows from all sheets in a workbook."""
    rows = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            if row and any(cell is not None and str(cell).strip() for cell in row):
                rows.append(','.join('' if cell is None else str(cell) for cell in row))
    return rows

def extract_xlsx(file_path):
    """Extract content from XLSX file"""
    if not XLSX_AVAILABLE:
        return None, "XLSX extraction library not available"

    try:
        rows = []
        # Try cached values first, then formula text if sheet appears empty.
        for data_only in (True, False):
            workbook = load_workbook(file_path, read_only=False, data_only=data_only)
            try:
                rows = _read_xlsx_rows(workbook)
            finally:
                workbook.close()
            if rows:
                break

        if not rows:
            return None, "No cell values found in XLSX file"
        return '\n'.join(rows), None
    except Exception as e:
        logger.error(f"XLSX extraction error: {str(e)}")
        return None, str(e)

def detect_file_extension(file_path):
    """Infer file extension from path suffix or file signature."""
    suffix = Path(file_path).suffix.lower()
    if suffix in EXTRACTION_FUNCTIONS:
        return suffix
    if is_xlsx_file(file_path):
        return '.xlsx'
    return None

def download_file(url, filename=None, content_type_hint=None):
    """
    Download file from URL to temporary location with size limits
    
    Args:
        url: URL to download from
        filename: Optional original filename from client (used when URL has no extension)
        content_type_hint: Optional MIME type from client
        
    Returns:
        tuple: (file_path, file_extension, error_message)
    """
    try:
        response = requests.get(
            url, 
            timeout=CONFIG['REQUEST_TIMEOUT'], 
            stream=True,
            allow_redirects=True
        )
        response.raise_for_status()
        
        # Check content length header
        content_length = response.headers.get('Content-Length')
        if content_length:
            try:
                size = int(content_length)
                if size > CONFIG['MAX_FILE_SIZE']:
                    return None, None, f"File too large. Maximum size: {CONFIG['MAX_FILE_SIZE'] / (1024*1024):.1f}MB"
            except ValueError:
                pass  # Invalid content-length, continue
        
        response_content_type = response.headers.get('Content-Type', '')
        file_extension = resolve_file_extension(filename, content_type_hint)

        if not file_extension:
            url_path = Path(url.split('?')[0])  # Remove query parameters
            if url_path.suffix:
                file_extension = url_path.suffix.lower()

        if not file_extension:
            file_extension = extension_from_content_type(response_content_type)
        
        # Create temporary file
        suffix = file_extension or '.tmp'
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        
        # Stream download with size check
        downloaded = 0
        try:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    downloaded += len(chunk)
                    if downloaded > CONFIG['MAX_FILE_SIZE']:
                        temp_file.close()
                        os.unlink(temp_file.name)
                        return None, None, f"File too large. Maximum size: {CONFIG['MAX_FILE_SIZE'] / (1024*1024):.1f}MB"
                    temp_file.write(chunk)
        finally:
            temp_file.close()
        
        if not file_extension:
            file_extension = detect_file_extension(temp_file.name)

        final_path = ensure_file_extension_path(temp_file.name, file_extension)
        logger.info(f"Downloaded file: {downloaded} bytes, extension: {file_extension}")
        return final_path, file_extension, None
        
    except requests.Timeout:
        return None, None, f"Request timeout (>{CONFIG['REQUEST_TIMEOUT']}s)"
    except requests.RequestException as e:
        logger.error(f"Download error: {str(e)}")
        return None, None, f"Failed to download file: {str(e)}"
    except Exception as e:
        logger.error(f"Unexpected download error: {str(e)}")
        return None, None, f"Unexpected error: {str(e)}"

def extension_from_content_type(content_type):
    """
    Map MIME type to file extension.

    Order matters: spreadsheet types must be checked before wordprocessing,
    because spreadsheet MIME types contain 'officedocument' (not 'document' alone).
    """
    if not content_type:
        return None

    content_type_lower = content_type.lower().split(';')[0].strip()

    if 'spreadsheetml' in content_type_lower or 'excel' in content_type_lower:
        return '.xlsx'
    if 'wordprocessingml' in content_type_lower:
        return '.docx'
    if content_type_lower in ('application/msword',) or 'ms-word' in content_type_lower:
        return '.doc'
    if 'pdf' in content_type_lower:
        return '.pdf'
    if 'csv' in content_type_lower or content_type_lower == 'text/csv':
        return '.csv'
    if content_type_lower == 'text/plain':
        return '.txt'

    return None

def resolve_file_extension(filename=None, content_type=None):
    """
    Resolve file extension from filename or Content-Type.
    
    Args:
        filename: Original file name (optional)
        content_type: MIME type string (optional)
        
    Returns:
        str | None: Lowercase extension including dot, or None
    """
    if filename:
        suffix = Path(filename).suffix.lower()
        if suffix:
            return suffix

    return extension_from_content_type(content_type)

# Extraction function mapping
EXTRACTION_FUNCTIONS = {
    '.pdf': extract_pdf,
    '.docx': extract_docx,
    '.doc': extract_doc,
    '.csv': extract_csv,
    '.txt': extract_txt,
    '.xlsx': extract_xlsx,
}

def try_extract_with_fallback(file_path, file_extension=None):
    """
    Try extraction with multiple methods if file extension is unknown
    
    Args:
        file_path: Path to the file
        file_extension: Known extension or None
        
    Returns:
        tuple: (content, detected_extension, error_message)
    """
    if not file_extension:
        file_extension = detect_file_extension(file_path)

    # If extension is known, use only the matching extractor (no cross-format fallback).
    if file_extension and file_extension in EXTRACTION_FUNCTIONS:
        extract_func = EXTRACTION_FUNCTIONS[file_extension]
        try:
            content, error = extract_func(file_path)
            if content and not error:
                return content, file_extension, None
            return None, file_extension, error or "Failed to extract content from file"
        except Exception as e:
            logger.debug(f"Extraction with {file_extension} failed: {str(e)}")
            return None, file_extension, str(e)
    
    # Try all supported formats
    for ext, extract_func in EXTRACTION_FUNCTIONS.items():
        if file_extension and ext == file_extension:
            continue  # Already tried
        try:
            content, error = extract_func(file_path)
            if content and not error:
                logger.info(f"Detected file type as {ext} via content analysis")
                return content, ext, None
        except Exception as e:
            logger.debug(f"Extraction with {ext} failed: {str(e)}")
            continue
    
    return None, file_extension, "Could not extract content with any supported method"

@app.route('/extract', methods=['POST', 'GET'])
@limiter.limit("30 per minute")
@require_api_key
def extract():
    """Extract content from file URL"""
    file_path = None
    try:
        # Get file URL and optional metadata from request
        if request.method == 'POST':
            data = request.get_json() or {}
            file_url = data.get('url') or request.form.get('url')
            filename = data.get('filename')
            content_type = data.get('contentType') or data.get('content_type')
        else:
            file_url = request.args.get('url')
            filename = request.args.get('filename')
            content_type = request.args.get('contentType') or request.args.get('content_type')

        if not file_url:
            logger.warning("Extraction request without URL")
            return jsonify({
                'error': 'Missing file URL. Provide "url" parameter in query string (GET) or JSON body (POST)'
            }), 400
        
        # Validate URL
        is_valid, error_msg = validate_url(file_url)
        if not is_valid:
            logger.warning(f"Invalid URL rejected: {file_url[:100]}")
            return jsonify({'error': f'Invalid URL: {error_msg}'}), 400
        
        logger.info(
            f"Extraction request for URL: {file_url[:100]}... "
            f"filename={filename}, content_type={content_type}"
        )

        # Download file
        file_path, file_extension, error = download_file(
            file_url,
            filename=filename,
            content_type_hint=content_type,
        )
        if error:
            logger.error(f"Download failed: {error}")
            return jsonify({'error': error}), 400
        
        # Extract content
        content, detected_ext, extract_error = try_extract_with_fallback(file_path, file_extension)
        
        if extract_error:
            logger.error(f"Extraction failed: {extract_error}")
            return jsonify({
                'error': f'Failed to extract content: {extract_error}',
                'file_type': detected_ext or file_extension
            }), 400
        
        if content is None:
            logger.warning(f"Could not extract content from file: {file_extension}")
            return jsonify({
                'error': 'Unsupported file type or failed to extract content',
                'file_type': detected_ext or file_extension,
                'supported_types': SUPPORTED_EXTENSIONS
            }), 400
        
        logger.info(f"Successfully extracted {detected_ext or file_extension} file, length: {len(content)}")
        return jsonify({
            'success': True,
            'content': content,
            'file_type': detected_ext or file_extension,
            'content_length': len(content)
        }), 200
            
    except Exception as e:
        logger.error(f"Unexpected error in extract endpoint: {str(e)}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500
    finally:
        # Clean up temporary file
        if file_path and os.path.exists(file_path):
            try:
                os.unlink(file_path)
            except Exception as e:
                logger.warning(f"Failed to delete temp file {file_path}: {str(e)}")

@app.route('/extract-base64', methods=['POST'])
@limiter.limit("30 per minute")
@require_api_key
def extract_base64():
    """Extract content from base64-encoded file data"""
    file_path = None
    try:
        data = request.get_json() or {}
        base64_input = data.get('base64')
        filename = data.get('filename')
        content_type = data.get('contentType')
        
        if not base64_input or not isinstance(base64_input, str):
            logger.warning("Extraction request without valid base64 payload")
            return jsonify({
                'error': 'Missing base64 data. Provide "base64" in JSON body.'
            }), 400
        
        if len(base64_input) > CONFIG['MAX_FILE_SIZE'] * 2:
            return jsonify({
                'error': f'Base64 payload too large. Maximum file size: {CONFIG["MAX_FILE_SIZE"] / (1024*1024):.1f}MB'
            }), 400
        
        # Support data URLs like: data:<mime>;base64,<payload>
        if ',' in base64_input and base64_input.strip().lower().startswith('data:'):
            try:
                header, payload = base64_input.split(',', 1)
                base64_input = payload.strip()
                if ';' in header and ':' in header:
                    detected_mime = header.split(':', 1)[1].split(';', 1)[0].strip()
                    if detected_mime and not content_type:
                        content_type = detected_mime
            except ValueError:
                return jsonify({'error': 'Invalid data URL format for base64 payload'}), 400
        
        try:
            file_bytes = base64.b64decode(base64_input, validate=True)
        except binascii.Error:
            return jsonify({'error': 'Invalid base64 data'}), 400
        
        if not file_bytes:
            return jsonify({'error': 'Decoded file is empty'}), 400
        
        if len(file_bytes) > CONFIG['MAX_FILE_SIZE']:
            return jsonify({
                'error': f'File too large. Maximum size: {CONFIG["MAX_FILE_SIZE"] / (1024*1024):.1f}MB'
            }), 400
        
        file_extension = resolve_file_extension(filename, content_type)
        if not file_extension and is_xlsx_bytes(file_bytes):
            file_extension = '.xlsx'
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_extension or '.tmp')
        try:
            temp_file.write(file_bytes)
        finally:
            temp_file.close()

        file_path = ensure_file_extension_path(
            temp_file.name,
            file_extension or detect_file_extension(temp_file.name),
        )
        file_extension = Path(file_path).suffix.lower() or file_extension
        logger.info(
            f"Base64 extraction request. filename={filename}, content_type={content_type}, "
            f"bytes={len(file_bytes)}, extension={file_extension}"
        )
        
        content, detected_ext, extract_error = try_extract_with_fallback(file_path, file_extension)
        
        if extract_error:
            logger.error(f"Base64 extraction failed: {extract_error}")
            return jsonify({
                'error': f'Failed to extract content: {extract_error}',
                'file_type': detected_ext or file_extension
            }), 400
        
        if content is None:
            return jsonify({
                'error': 'Unsupported file type or failed to extract content',
                'file_type': detected_ext or file_extension,
                'supported_types': SUPPORTED_EXTENSIONS
            }), 400
        
        return jsonify({
            'success': True,
            'content': content,
            'file_type': detected_ext or file_extension,
            'content_length': len(content)
        }), 200
    
    except Exception as e:
        logger.error(f"Unexpected error in extract-base64 endpoint: {str(e)}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500
    finally:
        if file_path and os.path.exists(file_path):
            try:
                os.unlink(file_path)
            except Exception as e:
                logger.warning(f"Failed to delete temp file {file_path}: {str(e)}")

@app.route('/test-celery', methods=['GET', 'POST'])
@require_api_key
def test_celery():
    """Enqueue a test Celery task to verify Redis + worker are connected."""
    try:
        from tasks import process_item
    except ImportError as exc:
        logger.error("Celery tasks not available: %s", exc)
        return jsonify({'error': 'Celery is not configured on this server'}), 503

    if request.method == 'POST':
        data = request.get_json() or {}
        item_id = data.get('id') or data.get('item_id')
    else:
        item_id = request.args.get('id') or request.args.get('item_id')

    if not item_id:
        item_id = 'test'

    try:
        async_result = process_item.delay(item_id)
    except Exception as exc:
        logger.error("Failed to enqueue Celery task: %s", exc, exc_info=True)
        return jsonify({
            'error': 'Failed to enqueue task. Check CELERY_BROKER_URL / REDIS_URL and Redis.',
            'details': str(exc),
        }), 503

    logger.info("Enqueued test Celery task id=%s item_id=%s", async_result.id, item_id)
    return jsonify({
        'success': True,
        'message': 'Task queued. Check worker logs for [process_item] output.',
        'task_id': async_result.id,
        'item_id': item_id,
        'status_url': f'/test-celery/{async_result.id}',
    }), 202


@app.route('/test-celery/<task_id>', methods=['GET'])
@require_api_key
def test_celery_status(task_id):
    """Poll result of a test Celery task by task id."""
    try:
        from celery_app import celery
        async_result = celery.AsyncResult(task_id)
    except Exception as exc:
        logger.error("Failed to read Celery task status: %s", exc)
        return jsonify({'error': 'Failed to read task status', 'details': str(exc)}), 503

    response = {
        'task_id': task_id,
        'state': async_result.state,
        'ready': async_result.ready(),
        'successful': async_result.successful() if async_result.ready() else None,
    }

    if async_result.failed():
        response['error'] = str(async_result.result)
    elif async_result.ready() and async_result.successful():
        response['result'] = async_result.result

    return jsonify(response), 200


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'pdf_support': PDF_AVAILABLE,
        'docx_support': DOCX_AVAILABLE,
        'doc_support': DOC_AVAILABLE,
        'xlsx_support': XLSX_AVAILABLE,
        'max_file_size_mb': CONFIG['MAX_FILE_SIZE'] / (1024 * 1024),
        'auth_required': bool(CONFIG.get('FILE_EXTRACTOR_KEY', ''))
    }), 200

@app.route('/', methods=['GET'])
def index():
    """Root endpoint with API information"""
    return jsonify({
        'message': 'File Extractor API',
        'endpoints': {
            '/extract': 'Extract content from file URL (GET or POST with url parameter) - Requires API key',
            '/extract-base64': 'Extract content from base64 file payload (POST with base64, optional filename/contentType) - Requires API key',
            '/test-celery': 'Enqueue test Celery task (GET/POST, optional id) - Requires API key',
            '/test-celery/<task_id>': 'Poll test Celery task status - Requires API key',
            '/health': 'Health check endpoint'
        },
        'supported_formats': SUPPORTED_EXTENSIONS,
        'max_file_size_mb': CONFIG['MAX_FILE_SIZE'] / (1024 * 1024),
        'authentication': 'API key required in Authorization header (Bearer <key> or just <key>)'
    }), 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
